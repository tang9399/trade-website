#!/usr/bin/env python3
"""
Excel → JSON 转换脚本
将运单Excel转成data/tracking.json，上传到GitHub后自动生效

使用方法:
  python3 excel_to_json.py <Excel文件路径> [输出JSON路径]

Excel格式:
  - Sheet "运单信息": 运单号 | 发货地 | 目的地 | 状态 | 状态文字 | 运输类型 | 重量 | 预计到达 | 发货地址
  - Sheet "物流轨迹": 运单号 | 时间 | 事件描述 | 是否已完成 | 是否当前节点
"""

import sys, json, os
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl -q')
    import openpyxl


def convert(excel_path, output_path="tracking.json"):
    wb = openpyxl.load_workbook(excel_path)
    
    # === Sheet 1: 运单信息 ===
    ws1 = wb["运单信息"]
    rows1 = list(ws1.iter_rows(min_row=2, values_only=True))
    
    # === Sheet 2: 物流轨迹 ===
    ws2 = wb["物流轨迹"]
    rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
    
    # 按运单号整理轨迹
    timeline_map = {}
    for row in rows2:
        no, time, text, done_str, current_str = row[0], row[1], row[2], row[3], row[4]
        if not no:
            continue
        no = str(no).strip().upper()
        if no not in timeline_map:
            timeline_map[no] = []
        timeline_map[no].append({
            "time": str(time) if time else "",
            "text": str(text) if text else "",
            "done": str(done_str).strip() in ("是", "yes", "Y", "1"),
            "current": str(current_str).strip() in ("是", "yes", "Y", "1"),
        })
    
    # 构建数据库
    db = OrderedDict()
    for row in rows1:
        no, from_city, to_city, status, status_text, trans_type, weight, eta, origin = (
            row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]
        )
        if not no:
            continue
        no = str(no).strip().upper()
        
        entry = {
            "trackingNo": no,
            "from": str(from_city).strip() if from_city else "",
            "to": str(to_city).strip() if to_city else "",
            "status": str(status).strip() if status else "transporting",
            "statusText": str(status_text).strip() if status_text else "运输中",
            "transportType": str(trans_type).strip() if trans_type else "铁路运输",
            "weight": str(weight).strip() if weight else "",
            "eta": str(eta).strip() if eta else "",
            "origin": str(origin).strip() if origin else "义乌福田街道涌金大道B9号",
            "timeline": timeline_map.get(no, [])
        }
        db[no] = entry
    
    # 写入JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 转换完成！共 {len(db)} 条运单")
    print(f"   输出文件: {output_path}")
    print(f"   上传此文件到 data/tracking.json 即可生效")


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        print("用法: python3 excel_to_json.py <Excel文件路径> [输出JSON路径]")
        print("示例: python3 excel_to_json.py tracking_template.xlsx tracking.json")
        sys.exit(1)
    
    excel = args[0]
    output = args[1] if len(args) > 1 else "tracking.json"
    convert(excel, output)
